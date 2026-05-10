"""Per-vehicle fuel-log export to .xlsx.

Replicates the structure of the original Vehicle_Fuel.xlsx workbook so a
Leon-style spreadsheet user can keep their existing pivot/chart muscle
memory:

    Sheet "Cost" with columns:
        Date | Milage | Amount L | Cost (<CUR>) | Petrol Station | Remarks
        | Distance | Consume (L/100 km) | Price/Liter | Price/Km
        | Days between | Column1

    Data rows carry live Excel formulas (=B3-B2, =(C3/G3)*100, …) so
    the workbook stays editable. A SUBTOTAL totals row at the bottom
    drives the Table totals (count / max / sum / average) — matching the
    original column-wise totalsRowFunction set.

The first data row has no previous odometer to subtract from, so its
Distance is hard-zero and Consume / Price-Km / Days-between are the
literal string "inf" — same convention as the original sheet so the
totals row's averages still ignore the row via SUBTOTAL.

Public entry point:
    build_vehicle_xlsx(vehicle_id) -> tuple[bytes, str]
        bytes: the workbook as an in-memory blob
        str:   suggested filename ("<safe_vehicle_name>_Fuel.xlsx")
"""

from __future__ import annotations

import io
import re
from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

import fuel


# ── Column layout ──────────────────────────────────────────────────────
# Order matters: columns are written A..L in this exact sequence.
_COLUMNS: list[tuple[str, str | None]] = [
    ("Date",                "count"),
    ("Milage",              "max"),
    ("Amount L",            "sum"),
    ("Cost ({cur})",        "sum"),
    ("Petrol Station",      None),
    ("Remarks",             None),
    ("Distance",            "sum"),
    ("Consume (L/100 km)",  "average"),
    ("Price/Liter",         "average"),
    ("Price/Km",            "average"),
    ("Days between",        "average"),
    ("Column1",             None),
]

# Width in Excel character units, matching the original Vehicle sheet so
# the file looks identical out of the box.
_COL_WIDTHS = {
    "A": 11.27,  "B": 18.27,  "C": 20.45,  "D": 18.36,
    "E": 18.54,  "F": 12.45,  "G": 18.45,  "H": 20.09,
    "I": 14.91,  "J": 13.73,  "K": 18.00,  "L": 8.43,
}


def _accounting_format(currency: str) -> tuple[str, str]:
    """Return (cost_format, price_per_liter_format) for the given currency.

    The original sheet uses a localized "TSh" prefix and Excel's standard
    accounting alignment (negatives in parens-style with leading minus,
    zeros shown as a single dash). For non-TZS vehicles we keep the same
    layout but swap the currency token.
    """
    cur = (currency or "TZS").upper()
    # TSh is the de-facto Tanzanian shilling abbreviation in Excel locales;
    # match the original file byte-for-byte for TZS vehicles.
    sym = "TSh" if cur == "TZS" else cur
    cost_fmt = (
        f'_-"{sym}"* #,##0.00_-;'
        f'\\-"{sym}"* #,##0.00_-;'
        f'_-"{sym}"* "-"??_-;'
        f'_-@_-'
    )
    price_fmt = (
        f'_-"{sym}"* #,##0_-;'
        f'\\-"{sym}"* #,##0_-;'
        f'_-"{sym}"* "-"??_-;'
        f'_-@_-'
    )
    return cost_fmt, price_fmt


def _safe_filename(name: str) -> str:
    """Slugify a vehicle name into a Windows/macOS-safe filename stem."""
    cleaned = re.sub(r"[^\w\-]+", "_", (name or "Vehicle").strip())
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned or "Vehicle"


def _parse_date(value: str) -> date | None:
    """Parse the YYYY-MM-DD date string used in fuel_log.csv."""
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def _to_float(value, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def build_vehicle_xlsx(vehicle_id: str) -> tuple[bytes, str]:
    """Build a per-vehicle fuel workbook in memory.

    Args:
        vehicle_id: The vehicle to export. Must exist in vehicles.csv.

    Returns:
        (xlsx_bytes, filename) — caller streams the bytes to the client
        and uses ``filename`` for Content-Disposition.

    Raises:
        ValueError: when the vehicle does not exist.
    """
    vehicles = fuel.load_vehicles()
    if vehicle_id not in vehicles:
        raise ValueError(f"Vehicle '{vehicle_id}' not found")
    vehicle = vehicles[vehicle_id]

    rows = [r for r in fuel.load_fuel_log() if r.get("vehicle_id") == vehicle_id]
    rows.sort(key=lambda r: (r.get("date", ""), r.get("fuel_id", "")))

    currency = (vehicle.get("currency") or "TZS").upper()
    cost_fmt, price_fmt = _accounting_format(currency)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cost"

    # ── Header row ──
    headers = [name.format(cur=currency) for name, _ in _COLUMNS]
    ws.append(headers)

    # ── Data rows ──
    # Row index in Excel terms (1-indexed). r=2 is the first data row.
    n_data = len(rows)
    for i, row in enumerate(rows):
        r = i + 2
        prev = r - 1  # previous data row, or the header row for i=0

        d = _parse_date(row.get("date", ""))
        odometer = _to_float(row.get("odometer_km"))
        liters = _to_float(row.get("liters"))
        cost = _to_float(row.get("total_cost"))
        station = row.get("station") or vehicle.get("default_payee", "")
        full_tank = str(row.get("full_tank", "")).strip().lower() == "true"
        remarks_default = "Full" if full_tank else "Partial"
        remarks = (row.get("remarks") or "").strip() or remarks_default

        if i == 0:
            # First entry: no previous odometer, so distance-derived
            # metrics fall back to the original sheet's "inf" sentinels.
            ws.cell(row=r, column=1, value=d).number_format = "mm-dd-yy"
            ws.cell(row=r, column=2, value=odometer).number_format = "0"
            ws.cell(row=r, column=3, value=liters)
            ws.cell(row=r, column=4, value=cost).number_format = cost_fmt
            ws.cell(row=r, column=5, value=station)
            ws.cell(row=r, column=6, value=remarks)
            ws.cell(row=r, column=7, value=0)
            ws.cell(row=r, column=8, value="inf")
            ws.cell(row=r, column=9, value=f"=D{r}/C{r}").number_format = price_fmt
            ws.cell(row=r, column=10, value="inf").number_format = cost_fmt
            ws.cell(row=r, column=11, value="inf")
            ws.cell(row=r, column=12, value="")
        else:
            ws.cell(row=r, column=1, value=d).number_format = "mm-dd-yy"
            ws.cell(row=r, column=2, value=odometer).number_format = "0"
            ws.cell(row=r, column=3, value=liters)
            ws.cell(row=r, column=4, value=cost).number_format = cost_fmt
            ws.cell(row=r, column=5, value=station)
            ws.cell(row=r, column=6, value=remarks)
            ws.cell(row=r, column=7, value=f"=B{r}-B{prev}")
            ws.cell(row=r, column=8, value=f"=(C{r}/G{r})*100")
            ws.cell(row=r, column=9, value=f"=D{r}/C{r}").number_format = price_fmt
            ws.cell(row=r, column=10, value=f"=D{r}/G{r}").number_format = cost_fmt
            ws.cell(row=r, column=11, value=f"=_xlfn.DAYS(A{r},A{prev})")
            ws.cell(row=r, column=12, value="")

    # ── Totals row ──
    # Always present, even when n_data == 0, to keep the SUBTOTAL formulas
    # documenting the column intent. Excel will show #DIV/0! for averages
    # over an empty Table — that is acceptable; the row anchors the layout.
    totals_row = max(n_data, 1) + 2  # one row past the last data row
    ws.cell(row=totals_row, column=1, value="=SUBTOTAL(103,Table1[Date])")
    ws.cell(row=totals_row, column=2, value="=SUBTOTAL(104,Table1[Milage])").number_format = "0"
    ws.cell(row=totals_row, column=3, value="=SUBTOTAL(109,Table1[Amount L])")
    cost_header = headers[3]
    ws.cell(row=totals_row, column=4, value=f"=SUBTOTAL(109,Table1[{cost_header}])").number_format = cost_fmt
    # E, F intentionally blank in the totals row — matches the original.
    ws.cell(row=totals_row, column=7, value="=SUBTOTAL(109,Table1[Distance])")
    ws.cell(row=totals_row, column=8, value="=SUBTOTAL(101,Table1[Consume (L/100 km)])")
    ws.cell(row=totals_row, column=9, value="=SUBTOTAL(101,Table1[Price/Liter])").number_format = price_fmt
    ws.cell(row=totals_row, column=10, value="=SUBTOTAL(101,Table1[Price/Km])").number_format = cost_fmt
    ws.cell(row=totals_row, column=11, value="=SUBTOTAL(101,Table1[Days between])")

    # ── Column widths ──
    for letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    # ── Excel Table object ──
    # openpyxl rejects a Table whose data area is zero rows. When the
    # vehicle has no fuel entries we skip the Table entirely and ship a
    # plain headers-only workbook so the download still works on an empty
    # vehicle (the totals row stays in place but without Table[…] refs we
    # rewrite its formulas to plain ranges).
    if n_data == 0:
        # Rewrite Table[…] refs to plain header-only ranges so Excel does
        # not throw on open. Header is row 1, the totals row is row 2, no
        # data rows in between.
        ws.cell(row=2, column=1, value=None)  # SUBTOTAL on empty range = 0
        ws.cell(row=2, column=2, value=None)
        ws.cell(row=2, column=3, value=None)
        ws.cell(row=2, column=4, value=None)
        ws.cell(row=2, column=7, value=None)
        ws.cell(row=2, column=8, value=None)
        ws.cell(row=2, column=9, value=None)
        ws.cell(row=2, column=10, value=None)
        ws.cell(row=2, column=11, value=None)
    else:
        last_col = get_column_letter(len(_COLUMNS))
        last_row = totals_row
        ref = f"A1:{last_col}{last_row}"
        table = Table(displayName="Table1", ref=ref)
        table.totalsRowCount = 1
        table.headerRowCount = 1
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        cols = []
        for idx, (name_template, totals_fn) in enumerate(_COLUMNS, start=1):
            col_name = name_template.format(cur=currency)
            tc = TableColumn(id=idx, name=col_name)
            if totals_fn:
                tc.totalsRowFunction = totals_fn
            cols.append(tc)
        table.tableColumns = cols
        ws.add_table(table)

    # ── Serialize ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{_safe_filename(vehicle.get('name', vehicle_id))}_Fuel.xlsx"
    return buf.read(), filename
