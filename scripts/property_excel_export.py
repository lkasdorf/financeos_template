"""Per-property Excel exporter — data-only multi-sheet workbook.

Used by /api/properties/excel. Generates a workbook with four sheets:

    Summary       — Property stammdaten + headline KPIs
    LUKU_Log      — Every electricity-token purchase for this property
    Water_Log     — Every water-bill payment for this property
    TX_Linked     — Every transactions.csv row referenced by the logs
                    (so the accountant can verify the books match)

Charts are NOT generated — the Dashboard owns visualization. The export
is a flat data hand-off: open in Excel, sort/filter/pivot if needed.

The function `build_property_xlsx(property_id)` returns (bytes, filename)
to match the contract that fuel_export.build_vehicle_xlsx uses, so the
serve.py handler treats them identically.
"""

from __future__ import annotations

import csv
import io
import sys
from datetime import date as _date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import utilities  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent
TX_CSV = REPO_ROOT / "data" / "transactions.csv"


def _slug(text: str) -> str:
    """Filesystem-safe slug — letters/digits/underscore only."""
    out = []
    for ch in text:
        if ch.isalnum():
            out.append(ch)
        elif ch in (" ", "-", "_"):
            out.append("_")
    return "".join(out).strip("_") or "property"


def _load_linked_tx(luku_rows: list[dict], water_rows: list[dict]) -> list[dict]:
    """Return the full TX rows from transactions.csv that are linked from
    the property's logs (via tx_import_id), in chronological order.
    """
    target_ids: set[str] = set()
    for r in luku_rows:
        if r.get("tx_import_id"):
            target_ids.add(r["tx_import_id"])
    for r in water_rows:
        if r.get("tx_import_id"):
            target_ids.add(r["tx_import_id"])
    if not target_ids:
        return []
    out: list[dict] = []
    with open(TX_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row.get("import_id") in target_ids:
                out.append(row)
    out.sort(key=lambda r: (r.get("date", ""), r.get("import_id", "")))
    return out


def build_property_xlsx(property_id: str) -> tuple[bytes, str]:
    """Build a per-property workbook in memory and return (bytes, filename).

    Raises:
        ValueError: when the property_id is unknown.
        ImportError: when openpyxl is not installed (handled by the API
            layer so the user gets a clear error instead of a 500).
    """
    properties = utilities.load_properties()
    if property_id not in properties:
        raise ValueError(f"Property '{property_id}' not found")
    prop = properties[property_id]

    luku_rows, water_rows = utilities.filter_logs_by_property(property_id)
    luku_rows.sort(key=lambda r: r.get("date", ""))
    water_rows.sort(key=lambda r: r.get("date", ""))
    kpis = utilities.property_kpis(luku_rows, water_rows)
    linked_tx = _load_linked_tx(luku_rows, water_rows)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for property Excel export"
        ) from exc

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")
    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    money_fmt = "#,##0.00"
    int_fmt = "#,##0"

    # ── Summary sheet ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Property — {prop.get('name', property_id)}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")

    summary_meta = [
        ("Property ID", prop.get("property_id", "")),
        ("Name", prop.get("name", "")),
        ("Address", prop.get("address", "")),
        ("Owner", prop.get("owner", "")),
        ("Default Account", prop.get("default_account", "")),
        ("Currency", prop.get("currency", "")),
        ("Electricity Payee", prop.get("electricity_payee", "")),
        ("Water Payee", prop.get("water_payee", "")),
        ("Electricity Meter", prop.get("electricity_meter", "")),
        ("Water Meter", prop.get("water_meter", "")),
        ("Water Control Number", prop.get("water_control_number", "")),
        ("Active", prop.get("active", "")),
    ]
    row = 3
    for label, value in summary_meta:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="KPIs").font = title_font
    row += 1
    kpi_pairs = [
        ("Total Strom (TZS)", kpis["total_strom"]),
        ("Total kWh", kpis["total_kwh"]),
        ("Total Water (TZS)", kpis["total_water"]),
        ("Avg TZS/kWh", kpis["avg_tzs_per_kwh"]),
        ("Strom YTD (TZS)", kpis["strom_ytd"]),
        ("kWh YTD", kpis["kwh_ytd"]),
        ("Water YTD (TZS)", kpis["wasser_ytd"]),
        ("Avg Strom / Monat (rolling 12m)", kpis["avg_strom_monthly"]),
        ("Avg Water / Monat (rolling 12m)", kpis["avg_wasser_monthly"]),
        ("Latest LUKU date", kpis["latest_luku_date"]),
        ("Latest Water date", kpis["latest_water_date"]),
        ("LUKU entries", kpis["luku_count"]),
        ("Water entries", kpis["water_count"]),
    ]
    for label, value in kpi_pairs:
        ws.cell(row=row, column=1, value=label).font = label_font
        cell = ws.cell(row=row, column=2, value=value)
        if isinstance(value, (int, float)) and "kWh" not in label and "entries" not in label:
            cell.number_format = money_fmt
        elif isinstance(value, (int, float)):
            cell.number_format = int_fmt if "entries" in label else "#,##0.00"
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    # ── LUKU_Log sheet ───────────────────────────────────────────────────
    luku_ws = wb.create_sheet("LUKU_Log")
    luku_headers = [
        "luku_id", "date", "property_id",
        "units_kwh", "total_price", "currency", "price_per_unit",
        "account", "meter", "tx_import_id", "note",
    ]
    _write_log_sheet(
        luku_ws, luku_headers, luku_rows,
        header_font=header_font, header_fill=header_fill,
        money_cols=("total_price", "price_per_unit"),
        kwh_cols=("units_kwh",),
        money_fmt=money_fmt,
    )

    # ── Water_Log sheet ──────────────────────────────────────────────────
    water_ws = wb.create_sheet("Water_Log")
    water_headers = [
        "water_id", "date", "property_id",
        "total_price", "currency",
        "control_number", "meter",
        "account", "tx_import_id", "note",
    ]
    _write_log_sheet(
        water_ws, water_headers, water_rows,
        header_font=header_font, header_fill=header_fill,
        money_cols=("total_price",),
        kwh_cols=(),
        money_fmt=money_fmt,
    )

    # ── TX_Linked sheet ──────────────────────────────────────────────────
    tx_ws = wb.create_sheet("TX_Linked")
    if linked_tx:
        tx_headers = list(linked_tx[0].keys())
    else:
        # Fallback header so an empty sheet still has a column row.
        tx_headers = [
            "import_id", "date", "account", "type", "amount", "currency",
            "payee", "category", "note", "tags",
        ]
    _write_log_sheet(
        tx_ws, tx_headers, linked_tx,
        header_font=header_font, header_fill=header_fill,
        money_cols=("amount", "transfer_to_amount"),
        kwh_cols=(),
        money_fmt=money_fmt,
    )

    # Auto-width per column on every sheet.
    for sheet in (luku_ws, water_ws, tx_ws):
        for col_idx, col_name in enumerate(_sheet_headers(sheet), start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = max(
                14, min(40, len(col_name) + 2)
            )

    # Serialize to bytes (in-memory; never touches disk).
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    today = _date.today().isoformat()
    name_slug = _slug(prop.get("name") or property_id)
    filename = f"{name_slug}_Utilities_{today}.xlsx"
    return data, filename


def _write_log_sheet(
    ws, headers: list[str], rows: list[dict],
    *, header_font, header_fill,
    money_cols: tuple[str, ...], kwh_cols: tuple[str, ...],
    money_fmt: str,
) -> None:
    """Render a header row + data rows on a worksheet.

    Numeric columns (money + kWh) are stored as numbers (not strings) so
    Excel can sort/sum them natively. Header row gets bold white text on
    a dark fill; the first row freezes for scrolling.
    """
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment_top_left()
    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, name in enumerate(headers, start=1):
            value = row.get(name, "")
            if name in money_cols and value not in ("", None):
                try:
                    value = float(value)
                except (TypeError, ValueError):
                    pass
            elif name in kwh_cols and value not in ("", None):
                try:
                    value = float(value)
                except (TypeError, ValueError):
                    pass
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if name in money_cols and isinstance(value, (int, float)):
                cell.number_format = money_fmt
            elif name in kwh_cols and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"


def _sheet_headers(ws) -> list[str]:
    """Return the header row of a worksheet (row 1) as a list of strings."""
    return [
        (cell.value or "") for cell in ws[1]
    ]


def Alignment_top_left():
    """Return an Alignment object with top-left anchoring.

    Wrapped in a function so the openpyxl import stays scoped to the
    main builder (which gracefully ImportError-raises if missing).
    """
    from openpyxl.styles import Alignment
    return Alignment(horizontal="left", vertical="top")


if __name__ == "__main__":
    # Quick CLI for ad-hoc generation outside the dashboard.
    if len(sys.argv) < 2:
        print("Usage: python scripts/property_excel_export.py <property_id> [out.xlsx]")
        sys.exit(2)
    pid = sys.argv[1]
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    data, filename = build_property_xlsx(pid)
    if out_path is None:
        out_path = Path(filename)
    out_path.write_bytes(data)
    print(f"Wrote {out_path} ({len(data):,} bytes)")
